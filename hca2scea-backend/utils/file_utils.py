"""
Import required modules.
"""
import argparse
import multiprocessing
from contextlib import contextmanager
import pandas as pd
import xml.etree.ElementTree as xm
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook

import utils.sra_utils
import utils.get_attribs

"""
Define functions.
"""
def split_list(accessions: [], n: int) -> []:
    """
    Function to split a list of SRA run accessions into a nested list containing smaller lists
    of SRA run accessions of length n. This is because the request will not work with lists of
    accessions approximately > 100 accessions in length. Returns the nested list.
    """
    parts_list = []
    for i in range(0, len(accessions), n):
        part = accessions[i:i + n]
        if part:
            parts_list.append(part)
    return parts_list

def check_file_type(fastq_map: {}) -> {}:
    """
    Function to check the type of data files returned.
    """
    for accession in fastq_map.keys():
        try:
            urls = fastq_map[accession]['urls']
            for i in range(0,len(urls)):
                fastq = 'fastq' in fastq_map[accession]['urls'][i]
                if fastq:
                    fastq_map[accession].update({'url%s' % (str(i+1)): urls[i],'type': 'fastq'})
                else:
                    bam = 'bam' in fastq_map[accession]['urls'][i]
                if bam:
                    fastq_map[accession].update({'url%s' % (str(i+1)): urls[i], 'type': 'bam'})
                else:
                    fastq_map[accession].update({'url%s' % (str(i+1)): urls[i], 'type': 'sra'})
            del fastq_map[accession]['urls']
        except:
            fastq_map[accession]['urls'] = None
    return fastq_map

def get_pubmed_metadata(project_pubmed_id: str,iteration: int) -> []:
    """
    Function to fetch publication metadata from an xml following a request to NCBI.
    A pubmed id is provided in the request url. If a project title or name is found but not publication title is found,
    a further request is sent to Europe PMC to try to find the publication title given the project title or project name.
    """
    xml_content = sra_utils.SraUtils.request_pubmed_metadata(project_pubmed_id)
    title,author_list,grant_list,article_doi_id = get_attribs.get_attributes_pubmed(xml_content,iteration)
    return title,author_list,grant_list,article_doi_id

def get_bioproject_metadata(bioproject_accession: str) -> []:
    """
    Function to fetch project metadata from an xml following a request to NCBI.
    An SRA Bioproject accession is provided in the request url.
    """
    xml_content = sra_utils.SraUtils.request_bioproject_metadata(bioproject_accession)
    project_name,project_title,project_description,project_pubmed_id = get_attribs.get_attributes_bioproject(xml_content,bioproject_accession)
    return project_name,project_title,project_description,project_pubmed_id

def get_experimental_metadata(accessions: [],accession_type: str) -> [[],str]:
    """
    Function to decide which accessions to send per request, if more than 1 is needed, to the SRA database
    via the request_info function. This decision is made based on the length of the accession list. If the accession list
    is >100 accessions in length, the list will first be split into a nested list of smaller
    lists of accessions and a nested list will be returned. Otherwise, an unnested list will be returned.
    """
    if len(accessions) < 100:
        xml = sra_utils.SraUtils.request_accession_info(accessions,accession_type=accession_type)
        size = 'small'
        return xml,size
    else:
        size = 'large'
        parts_list = split_list(accessions, n=100)
        xmls = []
        for p in range(0,len(parts_list)):
            xml = sra_utils.SraUtils.request_accession_info(parts_list[p],accession_type=accession_type)
            xmls.append(xml)
        return xmls,size

def fetch_experimental_metadata(accessions_list: [],accession_type: str) -> []:
    """
    Function to fetch metadata attributes associated with a list of either biosample or
    experiment accessions (biosample & experiment are accession types).
    """
    xml_content_result,size = get_experimental_metadata(accessions_list,accession_type=accession_type)
    if size == 'large':
        nested_list = []
        if accession_type == 'biosample':
            for xml_content in xml_content_result:
                nested_list.extend([get_attribs.get_attributes_biosample(element) for element in xml_content])
        elif accession_type == 'experiment':
            for xml_content in xml_content_result:
                for experiment_package in xml_content.findall('EXPERIMENT_PACKAGE'):
                    nested_list.extend([get_attribs.get_attributes_library_protocol(experiment_package)])
    else:
        if accession_type == 'biosample':
            nested_list = [get_attribs.get_attributes_biosample(element) for element in xml_content_result]
        elif accession_type == 'experiment':
            nested_list = []
            for experiment_package in xml_content_result.findall('EXPERIMENT_PACKAGE'):
                nested_list.extend([get_attribs.get_attributes_library_protocol(experiment_package)])
    return nested_list

@contextmanager
def poolcontext(*args, **kwargs):
    """
    Function for multiprocessing with number of threads, nthread (specified in optional arguments).
    """
    pool = multiprocessing.Pool(*args, **kwargs)
    yield pool
    pool.terminate()

def check_list_str(values: str) -> []:
    """
    Checks if an input accession list is a list of comma-separated strings (accessions). Returns a list
    of accessions if True.
    """
    if "," not in values:
        raise argparse.ArgumentTypeError("Argument list not valid: comma separated list required")
    return values.split(',')

def check_file(path: str) -> []:
    """
    Checks if an input file with a list of accessions is in the required format. The file should consist of a
    single column (list) of accessions with the column name "accession".
    """
    if not os.path.exists(path):
        raise argparse.ArgumentTypeError("file %s does not exist" % (path))
    try:
        df = pd.read_csv(path, sep="\t")
    except:
        raise argparse.ArgumentTypeError("file %s is not a valid format" % (path))
    try:
        geo_accession_list = list(df["accession"])
    except:
        raise argparse.ArgumentTypeError("accession list column not found in file %s" % (path))
    return geo_accession_list

def get_empty_df(workbook: object,tab_name: str) -> pd.DataFrame:
    """
    Initialise an empty dataframe for the tab name specified. The tab name is a tab expected to be
    found in the HCA metadata spreadsheet (e.g. Collection protocol tab or Donor organism tab).
    """
    sheet = workbook[tab_name]
    values = sheet.values
    empty_df = pd.DataFrame(values)
    cols = empty_df.loc[0,]
    tab = pd.DataFrame(columns=cols)
    return tab

def write_to_wb(workbook: Workbook, tab_name: str, tab_content: pd.DataFrame) -> None:
    """
    Write the dataframe (tab) to a tab in the active workbook.
    """
    worksheet = workbook[tab_name]
    row_not_filled = 6
    while True:
        if worksheet[f'A{row_not_filled}'].value or worksheet[f'B{row_not_filled}'].value:
            row_not_filled += 1
        else:
            break

    for index, key in enumerate(worksheet[4]):
        if not key.value:
            break
        if key.value not in tab_content.keys():
            continue
        for i in range(len(tab_content[key.value])):
            worksheet[f"{get_column_letter(index + 1)}{i + row_not_filled}"] = list(tab_content[key.value])[i]
